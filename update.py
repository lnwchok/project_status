from openpyxl import load_workbook, Workbook

# For Working Progress
def dump_data():
    folder = r"D:\\OneDrive - MPS\\BMC\\BMC-589\\03 PROJECT INFO\\094AZ1401-XX_MONTHLY REPORT\\data"
    filename = r"Progress.xlsx"
    sourcefile = folder+"\\"+filename
    wbs = load_workbook(filename=sourcefile, data_only=True, read_only=True)
    wss = wbs["Sum-Calc"]

    # months_cell = ["O4","Q4","S4","U4","W4","Y4","AA4","AC4","AE4","AG4"]
    range_cells = ["M","O","Q","S","U","W","Y","AA","AC","AE","AG"]
    # TODO Update please
    actual_cells=["N","P","R","T","V","X","Z","AB","AD","AF","AH"]
    progress = {
        "months" : [],
        "plan" : [],
        "actual" : []
    }

    for i in range_cells:
        progress["months"].append(wss[i+"87"].value)
        progress["plan"].append(wss[i+"90"].value)

    for j in actual_cells:
        progress["actual"].append(wss[j+"90"].value)

    return progress

def save2file():
    filename = r"D:\\OneDrive - MPS\\BMC\\BMC-589\\00 APP\\project_status\\data.xlsx"
    wb = Workbook()
    wb.create_sheet("progress")
    
    source = dump_data()
    len_h = len(source["months"])+1
    for h in range(1,len_h):
        wb["progress"].cell(row=h,column=1).value = source["months"][h-1]
        wb["progress"].cell(row=h,column=2).value = source["plan"][h-1]
        wb["progress"].cell(row=h,column=3).value = source["actual"][h-1]

    # print(row)
    wb.save(filename=filename)
    
save2file()




