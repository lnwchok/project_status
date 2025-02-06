from openpyxl import Workbook, load_workbook

# Please change normal SINGLE Backslash (\) to be DOUBLE Backslash (\\) for Path specify
# folder = r"D:\\OneDrive - MPS\\BMC\\BMC-589\\03 PROJECT INFO\\094AZ1401-XX_MONTHLY REPORT\\data"
# filename = r"Progress.xlsx"

# xlsxfile = folder+"\\"+filename

def getProgress(xlsxfile, sheetname):
    wb = load_workbook(filename=xlsxfile, data_only=True, read_only=True)
    ws = wb[sheetname]

    # months_cell = ["O4","Q4","S4","U4","W4","Y4","AA4","AC4","AE4","AG4"]
    range_cells = ["M","O","Q","S","U","W","Y","AA","AC","AE","AG"]
    # TODO Update please
    actual_cells=["N","P","R","T","V","X","Z","AB","AD","AF","AH"]
    progress = {
        "months" : [],
        "plan" : [],
        "actual" : []
    }

    for i in range_cells:
        progress["months"].append(ws[i+"87"].value)
        progress["plan"].append(ws[i+"90"].value)

    for j in actual_cells:
        progress["actual"].append(ws[j+"90"].value)

    return progress